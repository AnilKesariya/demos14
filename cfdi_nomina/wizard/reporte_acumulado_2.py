from odoo import models, fields
import base64
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook


class reporte_acumulado_2(models.TransientModel):
    _name = "cfdi_nomina.reporte.acumulado2"

    # period_id = fields.Many2one("account.period", string="Periodo")
    datas = fields.Binary("Excel Report")
    fname = fields.Char("Fname")
    slip_ids = fields.Many2many("hr.payslip", string=u"Payslip")
    employee_ids = fields.Many2many("hr.employee", string="Employees")
    rule_group_ids = fields.Many2many("hr.salary.rule.group", string="Groups of rules")

    def action_reporte(self, context=None):
        self._create_report()
        ids = self.env.context.get('active_id', False)
        this = self.browse(ids)
        return {
            'name': 'Reporte del periodo %s' % (this.period_id.name),
            'type': 'ir.actions.act_window',
            'res_model': "cfdi_nomina.reporte.acumulado2",
            'res_id': ids[0],
            'view_type': "form",
            'view_mode': 'form',
            'context': self.env.context,
            'target': 'new'
        }

    def _create_report(self):
        this = self.env.context.get('active_id', False)
        slip_line_obj = self.env['hr.payslip.line']
        # self.env.cr.execute("select line.id from hr_payslip_line line join hr_payslip slip on slip.id=line.slip_id where slip.period_id=%s and slip.state='done'", (this.period_id.id,))
        self.env.cr.execute("select line.id from hr_payslip_line line join hr_payslip slip on slip.id=line.slip_id where  slip.state='done'")
        line_ids = [x[0] for x in self.env.cr.fetchall()]
        data = {}
        structs = set()
        employee_ids = [x.id for x in this.employee_ids]
        slip_ids = [x.id for x in this.slip_ids]
        rule_group_ids = [x.id for x in this.rule_group_ids]
        all_slips = set()
        for line in slip_line_obj.browse(line_ids):
            if line.amount == 0:
                continue
            if employee_ids and line.slip_id.employee_id.id not in employee_ids:
                continue
            if slip_ids and line.slip_id.id not in slip_ids:
                continue
            if rule_group_ids and line.salary_rule_id.agrupacion.id not in rule_group_ids:
                continue
            if not rule_group_ids:
                code = line.code
                name = line.salary_rule_id.name
            else:
                code = line.salary_rule_id.agrupacion.name
                name = line.salary_rule_id.agrupacion.name        


    # def action_reporte(self, cr, uid, ids, context=None):
    #     self._create_report(cr, uid, ids, context=context)
    #     this = self.browse(cr, uid, ids[0])
    #     return {
    #         'name': 'Reporte del periodo %s'%(this.period_id.name),
    #         'type': 'ir.actions.act_window',
    #         'res_model': "cfdi_nomina.reporte.acumulado2",
    #         'res_id': ids[0],
    #         'view_type': "form",
    #         'view_mode': 'form',
    #         'context': context,
    #         'target': 'new'
    #     }


    # def _create_report(self, cr, uid, ids, context=None):
    #     this = self.browse(cr, uid, ids[0])
    #     slip_line_obj = self.pool.get("hr.payslip.line")
    #     # cr.execute("select line.id from hr_payslip_line line join hr_payslip slip on slip.id=line.slip_id where slip.period_id=%s and slip.state='done'", (this.period_id.id,))
    #     line_ids = [x[0] for x in cr.fetchall()]
    #     data = {}
    #     structs = set()
    #     employee_ids = [x.id for x in this.employee_ids]
    #     slip_ids = [x.id for x in this.slip_ids]
    #     rule_group_ids = [x.id for x in this.rule_group_ids]
    #     all_slips = set()
    #     for line in slip_line_obj.browse(cr, uid, line_ids):
    #         if line.amount == 0:
    #             continue
    #         if employee_ids and line.slip_id.employee_id.id not in employee_ids:
    #             continue
    #         if slip_ids and line.slip_id.id not in slip_ids:
    #             continue
    #         if rule_group_ids and line.salary_rule_id.agrupacion.id not in rule_group_ids:
    #             continue
    #         if not rule_group_ids:
    #             code = line.code
    #             name = line.salary_rule_id.name
    #         else:
    #             code = line.salary_rule_id.agrupacion.name
    #             name = line.salary_rule_id.agrupacion.name
    #         data.setdefault(code, {})
    #         data[code].setdefault("account", (line.salary_rule_id.account_debit.code or line.salary_rule_id.account_credit.code))
    #         data[code].setdefault("name",  name)
    #         data[code].setdefault("total",  0)
    #         data[code]["total"] += line.amount
    #         data[code].setdefault(line.slip_id.struct_id.name,  {})
    #         structs.add(line.slip_id.struct_id.name)
    #         data[code][line.slip_id.struct_id.name].setdefault("total", 0)
    #         data[code][line.slip_id.struct_id.name]["total"] += line.amount
    #         data[code][line.slip_id.struct_id.name].setdefault("slips", []).append(line.slip_id.number)
    #         all_slips.add(line.slip_id.number)

    #     wb = Workbook()
    #     ws = wb.active

    #     ws['A1'] = ','.join(list(all_slips))

    #     COL_ACCOUNT = 2
    #     COL_RULE = COL_ACCOUNT + 1
    #     COL_TOTAL = COL_ACCOUNT + 2
    #     COL_SLIPS = COL_ACCOUNT + 3
    #     def set_rule_header(row):
    #         ws.cell(column=COL_ACCOUNT, row=row, value="Cuenta")
    #         ws.cell(column=COL_RULE, row=row, value="Regla Salarial")
    #         ws.cell(column=COL_TOTAL, row=row, value="Total")

    #     row = 3

    #     for code in data:
    #         rule_data = data[code]
    #         set_rule_header(row)
    #         row += 1
    #         ws.cell(column=COL_ACCOUNT, row=row, value=rule_data["account"])
    #         ws.cell(column=COL_RULE, row=row, value=rule_data["name"])
    #         ws.cell(column=COL_TOTAL, row=row, value=rule_data["total"])

    #         row += 2
    #         for struct in structs:
    #             ws.cell(column=COL_RULE, row=row, value=struct)
    #             ws.cell(column=COL_TOTAL, row=row, value=rule_data[struct]["total"] if struct in rule_data else 0)
    #             if struct in rule_data:
    #                 ws.cell(column=COL_SLIPS, row=row, value=",".join(rule_data[struct].get("slips",[])))
    #             row += 1

    #         row += 2

    #     datas = base64.b64encode(save_virtual_workbook(wb))
    #     # self.write(cr, uid, [this.id], {'fname': "Reporte por estructura %s"%this.period_id.name + ".xlsx", 'datas': datas})
    #     return True
